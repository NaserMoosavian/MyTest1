# -*- coding: utf-8 -*-
"""
Created on Tue Feb 11 23:01:04 2020

@author: Naser Moosavian
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pandas import DataFrame
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.animation as animation
import tkinter as tk
from scipy import stats



Banker_Counter = []
Player_Counter = []
Tie_Counter = []
Global_Counter = []
Num_Counter =[0]
Number = 0
iteration = 1
x = []
BetValue =1
Alpha =1
xc=0
Action = 0
WL=[]
WL1=[]
WL2=[]
switch_value =0
Nvalue = 0
Bvalue = 0
aaa=0
counter =0
Wincounter = 0
ccounter = 0
Loscounter =0
ecounter =0
fcounter =0
bcounter =0
dcounter =0
tcounter = 0
xcounter =0
ycounter = 0
NumWincounter =0
profit = 0
profit1 = 0
profit2 = 0
profit3 = 0
profit4 = 0
profit5 =0
variable = 1
RRRR=0
YYYY=0
NumWinSeries=[0]
E2=[]
EM = []      
ProfitSeries = [0]
BetSeries = [0]


BetList1 = [1,1,1,2]
BetList2 = [2,2,2,4]



multiplayer = 1
Prob = 0.5
Alpha =1
Beta = 1
switch_value1 = 0

def save_info():
    return

def info_banker():
    global Banker_Counter
    global Global_Counter
    global Num_Counter
    global x,xc
    
  
    xc +=1
    x.append(xc)
    a = [1]
    b = "B"
    Num_Counter.extend(a) 
    Banker_Counter.extend(a) 
    Global_Counter.extend(b)
    info_print()
    simulaion_function()
    return 
  
def info_player():
    global Player_Counter
    global Global_Counter
    global Num_Counter
    global x,xc
    xc +=1
    x.append(xc)
    a = [-1]
    b = "P"
    Num_Counter.extend(a) 
    Player_Counter.extend(a) 
    Global_Counter.extend(b)
    info_print()
    simulaion_function()
    return 

def info_tie():
    global Tie_Counter
    global Global_Counter
    global Num_Counter
    global x,xc
    xc +=1
    x.append(xc)
    a = [0]
    b = "T"
    Num_Counter.extend(a) 
    Tie_Counter.extend(a) 
    Global_Counter.extend(b)
    info_print()
    simulaion_function()
    return 

def info_print():
    global Global_Counter
    T = tk.Message(text = str(Global_Counter),  width=300)
    T.pack()
    T.place(x=20,y=20)
    return

def reset_function():
    global Player_Counter,Banker_Counter,Tie_Counter,WL, WL1, WL2, NumWinSeries
    global x,xc, AveData, MaxData, MinData, iteration,Action, variable
    global Global_Counter,Num_Counter, Number, BetValue,Alpha, E2, R, EM
    global switch_value, ProfitSeries, NumWincounter, Wincounter, ecounter
    global Nvalue, profit, Bvalue, counter, aaa, Loscounter, ccounter, BetSeries
    Banker_Counter = []
    Player_Counter = []
    Tie_Counter = []
    Global_Counter = []
    Num_Counter =[]
    Number = 0
    iteration = 1
    x = []
    BetValue =1
    Alpha =1
    xc=0
    Action = 0
    WL=[]
    WL1=[]
    WL2=[]
    switch_value =0
    Nvalue = 0
    Bvalue = 0
    aaa=0
    counter =0
    Wincounter = 0
    ccounter = 0
    Loscounter =0
    ecounter =0
    NumWincounter =0
    profit = 0
    variable = 1
    R=1
    NumWinSeries=[0]
    E2=[]
    EM = []      
    ProfitSeries = [0]
    BetSeries = []

    return

def save_data_function():
    global History
    global Num_Counter
    if len(Num_Counter)>=69:
        T = tk.Label(text = "DataSet Saved", bg = 'white', fg = 'black', borderwidth=0, height=2, width=20)   
        T.pack()
        T.place(x=803,y=332)
        wb = load_workbook('Data_History.xlsx')
        AC = wb['Sheet2'].cell(1,1).value
        for row, val in enumerate(Num_Counter):
            if row+2<=70:
                wb['Sheet1'].cell(row+2,AC).value = val
        wb['Sheet2'].cell(1,1).value = wb['Sheet2'].cell(1,1).value+1
        wb.save('Data_History.xlsx')
    else:
        T = tk.Label(text = "Not Enough DataSet", bg = 'white', fg = 'black', borderwidth=0, height=2, width=20)   
        T.pack()
        T.place(x=803,y=332)

    return

def Strike_Betting ():
    global bcounter, BetValue, ecounter, profit,NumWinSeries
    
    BetList1 = [1,2,3,5,8,13,21,34]

    
    if ecounter == 0:
        BetValue = BetList1[0]
    elif ecounter == -2:  
        BetValue = BetList1[1]
    elif ecounter == -4:
        BetValue = BetList1[2]
    elif ecounter == -6:
        BetValue = BetList1[3]
    elif ecounter == -8:
        BetValue = BetList1[4]
    elif ecounter == -10:
        BetValue = BetList1[5]
    elif ecounter == -12:
        BetValue = BetList1[6]
    elif ecounter == -14:
        BetValue = BetList1[7]

    return 

def Counterstrike_Betting (): 
    global bcounter, BetValue, ecounter

    if switch_value==1:
        BetList2 = [1,1,2]
    elif switch_value==2:
        BetList2 = [1,2,3]
    elif switch_value==3:
        BetList2 = [2,3,5]
    elif switch_value==4:
        BetList2 = [3,5,8]
    elif switch_value==5:
        BetList2 = [5,8,13]
       
    if bcounter == 0:
        BetValue = BetList2[0]
    elif bcounter == -1:
        BetValue = BetList2[1]
    elif bcounter == -2:
        BetValue = BetList2[2]

    return 


def simulaion_function():
    global screen1, History, Num_Counter, Action, ListSize, bcounter, dcounter
    global ccounter, variable, switch_value, E1, ProfitSeries, BetSeries, BetValue
    global Number, iteration, Alpha, profit, ecounter, NumWinSeries, Wincounter, Loscounter
    global xcounter, RRRR, YYYY
    ListSize = len(Num_Counter)
               

    ## Results
    if Num_Counter[-1]==0 or Action == 0:
        WL.append("T")
        
    else:
        if Action == 1 and Num_Counter[-1] ==1:
            WL.append("W")
            
            NumWinSeries.extend([1])

            if switch_value == 0:
                ecounter +=2
                

            bcounter = 0
                
            
            if profit==0:
                
                xcounter += 1
                
            profit += BetValue
            

            
        if Action == -1 and Num_Counter[-1] ==-1:
            WL.append("W")
        
                
            NumWinSeries.extend([1])

            if switch_value == 0:
                ecounter +=2

                
            bcounter = 0
                

            
            if profit==0:
                
                xcounter += 1
            
            profit += BetValue

        if Action == -1 and Num_Counter[-1] ==1:
            WL.append("L")
                
            NumWinSeries.extend([-1])

            if switch_value == 0:
                ecounter -=1
            
            if switch_value >= 1:
                bcounter -=1
                

            profit -= BetValue

        if Action == 1 and Num_Counter[-1] ==-1:
            WL.append("L")

            NumWinSeries.extend([-1])

            if switch_value == 0:
                ecounter -=1
            
            if switch_value >= 1:
                bcounter -=1
                

            profit -= BetValue
   
        
    T = tk.Message(text = str(WL),  bg = 'white', width=690)
    T.pack()
    T.place(x=1,y=153)
    
    
   
    if ecounter>0:
        ecounter = 0
        
    if  bcounter==-3:
        ecounter -=1
        switch_value = 0
        bcounter = 0

    if ecounter==-1:
        switch_value = 1
    elif ecounter==-3:
        switch_value = 2
    elif ecounter==-5:
        switch_value = 3
    elif ecounter==-7:
        switch_value = 4
    elif ecounter==-9:
        switch_value = 5
    else:
        switch_value = 0

        



    if switch_value ==0:
        Strike_Betting ()
    else:
        Counterstrike_Betting ()
        

            
    if profit >0: 
        profit = 0
        BetValue = 1
        Alpha =1
        

           
    if profit <=-30*multiplayer:
        #BetValue = 0
        Alpha =2
        

    
    BetValue *= Alpha
    
    ###########################################################################
    # Version9
    # Selection Strategy
    # Strategy 2 (Last 3 underdog)
    if len(Num_Counter)>=3:
        if (Num_Counter[-3]+Num_Counter[-2]+Num_Counter[-1]) >=0:
            Action = -1
        else:
            Action = 1
    

    if len(Num_Counter)>=5: 
        
        if Num_Counter[-5]==1 and Num_Counter[-4]==-1 and Num_Counter[-3]==-1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==-1:
            Action =1
        if Num_Counter[-5]==-1 and Num_Counter[-4]==1 and Num_Counter[-3]==1 and Num_Counter[-2]==1 and Num_Counter[-1] ==1:
            Action =-1
        
        if Num_Counter[-5]==-1 and Num_Counter[-4]==-1 and Num_Counter[-3]==-1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==-1:
            Action =-1
        if Num_Counter[-5]==1 and Num_Counter[-4]==1 and Num_Counter[-3]==1 and Num_Counter[-2]==1 and Num_Counter[-1] ==1:
            Action =1
            
        if Num_Counter[-4]==1 and Num_Counter[-3]==-1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==-1:
            Action =1

        if Num_Counter[-4]==-1 and Num_Counter[-3]==1 and Num_Counter[-2]==1 and Num_Counter[-1] ==1:
            Action =-1

        if Num_Counter[-5]==1 and Num_Counter[-4]==1 and Num_Counter[-3]==1 and Num_Counter[-2]==1 and Num_Counter[-1] ==-1:
            Action =-1

        if Num_Counter[-5]==-1 and Num_Counter[-4]==-1 and Num_Counter[-3]==-1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==1:
            Action =1

    if len(Num_Counter)>=6: 
        if Num_Counter[-6]==1 and Num_Counter[-5]==1 and Num_Counter[-4]==1 and Num_Counter[-3]==1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==1:
            Action =-1

        if Num_Counter[-6]==-1 and Num_Counter[-5]==-1 and Num_Counter[-4]==-1 and Num_Counter[-3]==-1 and Num_Counter[-2]==1 and Num_Counter[-1] ==-1:
            Action =1

            
    if len(Num_Counter)>=7: 
        if Num_Counter[-7]==1 and Num_Counter[-6]==1 and Num_Counter[-5]==1 and Num_Counter[-4]==1 and Num_Counter[-3]==-1 and Num_Counter[-2]==1 and Num_Counter[-1] ==1:
            Action =-1

        if Num_Counter[-7]==-1 and Num_Counter[-6]==-1 and Num_Counter[-5]==-1 and Num_Counter[-4]==-1 and Num_Counter[-3]==1 and Num_Counter[-2]==-1 and Num_Counter[-1] ==-1:
            Action =1

        
    ProfitSeries.extend([profit])
    BetSeries.extend([BetValue])        
    
    
    if Action>0:
        w = tk.Label(screen1, text = "BANKER", bg = 'red', fg = 'white', height = '3', width = '15')
    elif Action<0:
        w = tk.Label(screen1, text = "PLAYER", bg = 'blue', fg = 'white', height = '3', width = '15')
    else:
        w = tk.Label(screen1, text = "WAIT", bg = 'white', fg = 'black', height = '3', width = '15')

    w.pack()
    w.place(x=810,y=125)
    w.config(font = ("arial", 12))

    B1 = np.round(BetValue)


    w1 = tk.Label(screen1, text = " CLR Bet  " + str(B1), bg = 'white', fg = 'black', height = '2', width = '15')
    w1.pack()
    w1.place(x=810,y=195)
    w1.config(font = ("arial", 12))
    
    
    w1 = tk.Label(screen1, text = " Profit  " + str(profit), bg = 'white', fg = 'black', height = '2', width = '15')
    w1.pack()
    w1.place(x=810,y=330)
    w1.config(font = ("arial", 12))

    w1 = tk.Label(screen1, text = " Max Budget  " + str(min(ProfitSeries)-BetValue), bg = 'white', fg = 'black', height = '2', width = '15')
    w1.pack()
    w1.place(x=810,y=380)
    w1.config(font = ("arial", 12))
    
    w1 = tk.Label(screen1, text = " Net Prof  " +str(xcounter), bg = 'white', fg = 'black', height = '2', width = '15')
    w1.pack()
    w1.place(x=810,y=480)
    w1.config(font = ("arial", 12))
    
    w1 = tk.Label(screen1, text = " Min Prof  " + str(min(ProfitSeries)), bg = 'white', fg = 'black', height = '2', width = '15')
    w1.pack()
    w1.place(x=810,y=430)
    w1.config(font = ("arial", 12))
    
    #save_data.place(x=800,y=330)
    
    return

def animate(i):
    global Num_Counter,x
    global ax1, ax2

    if len(ProfitSeries)>=1:
        fig1 = ProfitSeries
    else:
        fig1 = [-1,0,1,0,-1,0,1]
        
    ax1.clear()   
    df = DataFrame(fig1)
    df.plot(ax=ax1, color = 'blue', fontsize=0.01)

    ax1.get_legend().remove()
    ###########################################################################
    
    if len(BetSeries)>=1:
        fig2 = BetSeries
    else:
        fig2 = [-1,0,1,0,-1,0,1]
        
    ax2.clear()

    tf = DataFrame(fig2)
    tf.plot(ax=ax2, color = 'blue',  fontsize=0.01)

    ax2.get_legend().remove()
    
    return

def figure_function(screen1):
    global ax1, fig, line1
    fig = plt.figure(figsize=(4,3), dpi=100)
    fig.subplots_adjust()
    ax1 = fig.add_subplot(111)
    line1 = FigureCanvasTkAgg(fig, screen1)
    line1.get_tk_widget().pack(side=tk.LEFT)
    #ani = animation.FuncAnimation(fig, animate, interval=1000)
    plt.show()
    return

def figure_function2(screen1):
    global ax2
    fig2 = plt.figure(figsize=(4,3), dpi=100)
    fig2.subplots_adjust()
    ax2 = fig2.add_subplot(111)
    line2 = FigureCanvasTkAgg(fig2, screen1)
    line2.get_tk_widget().pack(side=tk.LEFT)
    #ani = animation.FuncAnimation(fig2, animate, interval=1000)
    plt.show()
    return


    
def BAC_CALL():
    global Global_Counter, screen1, Num_Counter
    screen1 = tk.Tk()
    screen1.geometry("1000x550")
    screen1.title("BACCARAT")
    register_banker = tk.Button(text = "BANKER", bg = 'red', fg = 'white', width = '20', height = '4', command = info_banker)
    register_banker.place(x=800,y=40)
    register_player = tk.Button(text = "PLAYER", bg = 'blue', fg = 'white', width = '20', height = '4', command = info_player)
    register_player.place(x=600,y=40)
    register_tie = tk.Button(text = "B6", bg = 'green', fg = 'white',width = '5', height = '2', command = info_tie)
    register_tie.place(x=753,y=40)
    figure_function(screen1)
    figure_function2(screen1)

    
    #print("NEW DATA")


    tk.mainloop()
    return
    

A=BAC_CALL()
